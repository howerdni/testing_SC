import streamlit as st
import pandas as pd
import io
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import re

class SCCalculator:
    def __init__(self):
        # Streamlit page configuration
        st.set_page_config(page_title="短路电流计算结果筛选器", layout="wide")
        st.title("短路电流计算结果筛选器")

        # Initialize session state
        if 'result_dfs' not in st.session_state:
            st.session_state.result_dfs = {}
        if 'bus_names' not in st.session_state:
            st.session_state.bus_names = []
        if 'files_loaded' not in st.session_state:
            st.session_state.files_loaded = False
        if 'ds_input' not in st.session_state:
            st.session_state.ds_input = ""
        if 'ds1_input' not in st.session_state:
            st.session_state.ds1_input = ""
        if 'uploaded_files' not in st.session_state:
            st.session_state.uploaded_files = []
        if 'selected_bus' not in st.session_state:
            st.session_state.selected_bus = ""

        # File uploader
        st.subheader("上传CSV文件")
        uploaded_files = st.file_uploader("选择CSV文件", type=["csv"], accept_multiple_files=True)

        # Handle file removal or new uploads
        if not uploaded_files and st.session_state.uploaded_files:
            # Files were removed
            st.session_state.files_loaded = False
            st.session_state.uploaded_files = []
            st.session_state.bus_names = []
            st.session_state.result_dfs = {}
            st.session_state.selected_bus = ""
            st.info("已移除所有文件，请上传新文件以继续。")
        elif uploaded_files and uploaded_files != st.session_state.uploaded_files:
            # New or different files uploaded
            self.load_files(uploaded_files)

        # DS and DS1 inputs (shown regardless of files_loaded)
        st.subheader("输入参数")
        col1, col2 = st.columns(2)

        with col1:
            st.write("母线名 (DS, 逗号分隔):")
            st.text_input("DS输入", value=st.session_state.ds_input, key="ds_input_field")
            st.caption("可用中文逗号（，）或英文逗号（,）分隔。提示：使用 Ctrl+A 或 Cmd+A 选择全部文本删除。")
            
            # Clear DS input button
            if st.button("清除DS输入", key="clear_ds_button"):
                st.session_state.ds_input = ""
                st.rerun()

            # Bus name selection (only shown if files are loaded)
            if st.session_state.files_loaded:
                st.write("选择母线名以追加到DS输入:")
                selected_bus = st.selectbox(
                    "选择母线名",
                    [""] + st.session_state.bus_names,
                    key="ds_suggest",
                    index=0
                )

                # Function to handle appending
                def append_to_ds():
                    if selected_bus:
                        # Use the current text_input value
                        current_ds = st.session_state.get("ds_input_field", "").strip()
                        new_ds = selected_bus if not current_ds else f"{current_ds}，{selected_bus}"
                        st.session_state.ds_input = new_ds
                        st.write(f"已追加: {selected_bus}，当前DS输入: {new_ds}")
                        st.session_state.selected_bus = ""  # Reset selection
                        st.rerun()  # Force UI refresh
                    else:
                        st.warning("请先从下拉菜单选择一个母线名")

                # Button to append selected bus name
                st.button("追加到DS", key="append_ds_button", on_click=append_to_ds)
                
                # JavaScript to handle Enter key press
                st.markdown("""
                    <script>
                    document.addEventListener('DOMContentLoaded', function() {
                        const selectBox = document.querySelector('select[data-testid="stSelectbox"]');
                        if (selectBox) {
                            selectBox.addEventListener('keydown', function(event) {
                                if (event.key === 'Enter' && this.value !== '') {
                                    const button = document.querySelector('button[kind="secondary"]');
                                    if (button) {
                                        button.click();
                                    }
                                }
                            });
                        }
                    });
                    </script>
                """, unsafe_allow_html=True)

        with col2:
            st.write("显示名称 (DS1, 逗号分隔):")
            ds1_input = st.text_input("DS1输入", value=st.session_state.ds1_input, key="ds1_input_field")
            st.caption("可用中文逗号（，）或英文逗号（,）分隔")

            # Store DS1 input
            st.session_state.ds1_input = ds1_input
            self.ds1_input = ds1_input
            self.uploaded_files = uploaded_files

        # Calculate button
        if st.button("计算"):
            self.calculate()

        # Display results
        if st.session_state.result_dfs:
            st.subheader("计算结果")
            for file_name, df in st.session_state.result_dfs.items():
                with st.expander(f"结果: {file_name}"):
                    st.dataframe(df, use_container_width=True)

        # Export button
        if st.session_state.result_dfs:
            excel_data = self.export_to_excel()
            st.download_button(
                label="导出到Excel",
                data=excel_data,
                file_name="short_circuit_results.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

    def load_files(self, uploaded_files):
        """Load CSV files and extract unique bus names."""
        with st.spinner("正在加载文件..."):
            bus_names = set()
            for uploaded_file in uploaded_files:
                try:
                    df = pd.read_csv(uploaded_file, encoding='gbk', index_col=False)
                    if '母线名' in df.columns:
                        bus_names.update(df['母线名'].dropna().astype(str).unique())
                    else:
                        st.warning(f"文件 {uploaded_file.name} 缺少 '母线名' 列")
                except Exception as e:
                    st.error(f"加载文件 {uploaded_file.name} 失败: {str(e)}")
                    return
            st.session_state.bus_names = sorted(list(bus_names))
            st.session_state.files_loaded = True
            st.session_state.uploaded_files = uploaded_files
            st.session_state.result_dfs = {}
            st.session_state.selected_bus = ""
            st.success("文件加载完成！请在下方输入DS和DS1。")

    def calculate(self):
        if not st.session_state.uploaded_files:
            st.error("请先上传CSV文件")
            return

        # Split DS and DS1 inputs using both English and Chinese commas
        ds = [x.strip() for x in re.split(r'[,\uFF0C]', st.session_state.ds_input) if x.strip()]
        ds1 = [x.strip() for x in re.split(r'[,\uFF0C]', st.session_state.ds1_input) if x.strip()]

        if not ds or not ds1:
            st.error("请填写DS和DS1")
            return

        if len(ds) != len(ds1):
            st.error("DS和DS1的条目数量必须相同")
            return

        st.session_state.result_dfs.clear()

        for uploaded_file in st.session_state.uploaded_files:
            file_name = uploaded_file.name
            try:
                uploaded_file.seek(0)
                sccp = pd.read_csv(uploaded_file, encoding='gbk', index_col=False)
                required_columns = ['母线名', '故障类型', '基电压']
                if not all(col in sccp.columns for col in required_columns):
                    missing = [col for col in required_columns if col not in sccp.columns]
                    st.error(f"文件 {file_name} 缺少必要列: {', '.join(missing)}")
                    return

                if len(sccp.columns) < 5:
                    st.error(f"文件 {file_name} 列数不足，缺少短路电流数据（第5列）")
                    return

                S2 = []
                S1 = []

                for i in ds:
                    found = False
                    for row in sccp.itertuples():
                        if i in row.母线名:
                            found = True
                            if row.故障类型 == '单相':
                                dict_sccp = {
                                    '母线名': row.母线名,
                                    'sc': row[5],
                                    '基电压': getattr(row, '基电压', '-')
                                }
                                S1.append(dict_sccp)
                            elif row.故障类型 == '三相':
                                dict_sccp = {
                                    '母线名': row.母线名,
                                    'sc': row[5],
                                    '基电压': getattr(row, '基电压', '-')
                                }
                                S2.append(dict_sccp)
                    if not found:
                        st.warning(f"文件 {file_name} 中未找到母线名包含 '{i}' 的记录")

                if not S1 and not S2:
                    st.error(f"文件 {file_name} 未找到任何匹配的单相或三相故障数据")
                    return

                substation2 = []
                sc2 = []
                base_voltage2 = []
                for i in S2:
                    substation2.append(i['母线名'])
                    sc2.append(i['sc'])
                    base_voltage2.append(i['基电压'])

                SD2 = {'substation': substation2, 'sc': sc2, 'base_voltage': base_voltage2}
                df2 = pd.DataFrame(SD2)

                substation1 = []
                sc1 = []
                base_voltage1 = []
                for i in S1:
                    substation1.append(i['母线名'])
                    sc1.append(i['sc'])
                    base_voltage1.append(i['基电压'])

                SD1 = {'substation': substation1, 'sc': sc1, 'base_voltage': base_voltage1}
                df1 = pd.DataFrame(SD1)

                if df2.empty and df1.empty:
                    st.error(f"文件 {file_name} 处理后未生成有效数据，请检查DS输入和CSV内容")
                    return

                X1 = list(zip(ds, ds1))
                df2c = df2.copy()
                df1c = df1.copy()
                DF2 = pd.DataFrame()
                DF1 = pd.DataFrame()

                for i in df2.index:
                    matched = False
                    for name in X1:
                        if df2.loc[i]['substation'] == name[0]:
                            df2c.at[i, 'sub_name'] = name[1]
                            matched = True
                            break
                    if not matched:
                        df2c.at[i, 'sub_name'] = df2.loc[i]['substation']

                DF2['sub_name'] = df2c['sub_name']
                DF2['sc'] = df2c['sc']
                DF2['base_voltage'] = df2c['base_voltage']

                for i in df1.index:
                    matched = False
                    for name in X1:
                        if df1.loc[i]['substation'] == name[0]:
                            df1c.at[i, 'sub_name'] = name[1]
                            matched = True
                            break
                    if not matched:
                        df1c.at[i, 'sub_name'] = df1.loc[i]['substation']

                DF1['sub_name'] = df1c['sub_name']
                DF1['sc'] = df1c['sc']
                DF1['base_voltage'] = df1c['base_voltage']

                result_df = pd.DataFrame()
                result_df['sub_name'] = DF2['sub_name']
                result_df['基电压'] = DF2['base_voltage']
                result_df['三相'] = DF2['sc']
                result_df['单相'] = DF1['sc']

                result_df = result_df.fillna('-')
                result_df[['三相', '单相']] = result_df[['三相', '单相']].apply(pd.to_numeric, errors='coerce').round(1)

                st.session_state.result_dfs[file_name] = result_df

            except Exception as e:
                st.error(f"处理文件 {file_name} 时发生错误: {str(e)}")
                return

        st.success("所有文件计算完成！")

    def export_to_excel(self):
        if not st.session_state.result_dfs:
            st.error("没有可导出的结果")
            return None

        output = io.BytesIO()
        wb = Workbook()
        wb.remove(wb.active)

        for file_name, df in st.session_state.result_dfs.items():
            ws = wb.create_sheet(title=file_name)
            ws['A1'] = file_name
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 2):
                for c_idx, value in enumerate(row, 1):
                    ws.cell(row=r_idx, column=c_idx).value = value

        wb.save(output)
        return output.getvalue()

if __name__ == "__main__":
    app = SCCalculator()
